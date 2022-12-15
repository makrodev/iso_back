import calendar
import requests
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

from datetime import date

from api import models as api_models
from main import settings as main_settings


class ExcelHelpers:
    def sheet_title(self, ws, cell, row):
        font = Font(
            bold=True,
        )

        side = Side(border_style="thick", color="000000")
        border = Border(
            top=side, bottom=side, left=side, right=side,
        )

        fill = PatternFill("solid", fgColor="00D3D3D3")
        align = Alignment(vertical='center', horizontal='center', wrapText=True)

        for key in cell:
            ws[f'{key}{row}'] = cell[key]
            ws[f'{key}{row}'].font = font
            ws[f'{key}{row}'].border = border
            ws[f'{key}{row}'].fill = fill
            ws[f'{key}{row}'].alignment = align

    def sheet_text(self, ws, cell, row, align_horizontal='left'):
        font = Font(
            bold=True,
        )

        align = Alignment(vertical='center', horizontal=align_horizontal, wrapText=True)

        for key in cell:
            ws[f'{key}{row}'] = cell[key]
            ws[f'{key}{row}'].font = font
            ws[f'{key}{row}'].alignment = align

    def get_max_width(self, cell, width):
        for cell__key in cell:
            str_len = len(str(cell[cell__key]))
            if str_len >= width:
                width = str_len
        return width

    def sheet_width(self, ws, cell, width):
        for key in cell:
            ws.column_dimensions[key].width = width

    def generate_next_month(self, year, month):
        year2 = year
        if month < 12:
            month2 = int(month) + 1
        else:
            month2 = 1
            year2 = int(year) + 1

        result = {
            "from": f"{year}-{month}-1 00:00:00.000000",
            "to": f"{year2}-{month2}-1 00:00:00.000000",
        }
        return result

    def generate_week(self, year, month, day_from, day_to):
        result = {
            "from": f"{year}-{month}-{day_from} 00:00:00.000000",
            "to": f"{year}-{month}-{day_to} 00:00:00.000000",
        }
        return result

    def percent_count(self, peace, all):
        if all != 0:
            return int(peace / all * 100)
        return 0

    def is_no_violation_text(self, is_no_violation_number):
        result_text = "Нет"
        if is_no_violation_number == 1:
            result_text = "Да"
        return result_text

    def datetime_to_text(self, datetime):
        return str(datetime).split(" ")[0]

    def days_in_month_func(self, year, month):
        days_in_month_list = []
        days_in_week_list = []
        last_day_of_month = calendar.monthrange(year, month)[1]
        for day in range(1, last_day_of_month + 1):
            intDay = date(year=year, month=month, day=day).weekday()
            if intDay in [0, 1, 2, 3, 4]:
                days_in_week_list.append(day)
                if day == last_day_of_month:
                    days_in_month_list.append(days_in_week_list)
            else:
                if days_in_week_list:
                    days_in_month_list.append(days_in_week_list)
                days_in_week_list = []
        return days_in_month_list

    def violation_weeks_count(self, year, month, days_in_month_list, process_id=None):
        result_list = []
        for i in range(0, 5):
            try:
                created_at_week_dict = ExcelHelpers.generate_week(
                    self, year=year, month=month,
                    day_from=days_in_month_list[i][0], day_to=days_in_month_list[i][-1],
                )
                if process_id == None:
                    violation_week_count = api_models.Violation.objects.filter(
                        created_at__gte=created_at_week_dict['from'],
                        created_at__lte=created_at_week_dict['to'],
                    ).count()
                else:
                    violation_week_count = api_models.Violation.objects.filter(
                        created_at__gte=created_at_week_dict['from'],
                        created_at__lte=created_at_week_dict['to'],
                        process_id=process_id,
                    ).count()
                result_list.append(violation_week_count)
            except IndexError:
                pass
        return result_list


class ViolationHelpers:
    def update(self, instance, validated_data):
        instance.client = validated_data.get('client', instance.client)
        instance.region = validated_data.get('region', instance.region)
        instance.shop = validated_data.get('shop', instance.shop)
        instance.department = validated_data.get('department', instance.department)
        instance.problem = validated_data.get('problem', instance.problem)
        instance.disparity = validated_data.get('disparity', instance.disparity)
        instance.comment = validated_data.get('comment', instance.comment)
        instance.photo = validated_data.get('photo', instance.photo)
        instance.response_admin = validated_data.get('response_admin', instance.response_admin)
        instance.response_person_description = validated_data.get('response_person_description', instance.response_person_description)
        instance.status = validated_data.get('status', instance.status)
        instance.process = validated_data.get('process', instance.process)
        instance.is_active = validated_data.get('is_active', instance.is_active)
        instance.save()
        return instance

    def send_bot_violation_message(self, instance):
        # BOT SETTINGS
        command_violation = api_models.Content.objects.get(key="commands_violation")

        bot_message = f"Вам пришло новое сообщение о нарушении\nНажмите сюда {command_violation.title}{instance.pk} чтобы ответить на него"

        base_url = f'{main_settings.TELEGRAM_DOMAIN}/bot{main_settings.BOT_KEY}/sendMessage?chat_id={instance.response_admin.tg_id}&text={bot_message}'
        requests.get(url=base_url)
